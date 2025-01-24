import openpyxl


def ReadData(File, Sheet, col, Row):
    worksheet = openpyxl.load_workbook(File)
    SheetName = worksheet[Sheet]
    return SheetName.cell(row=Row, column=col).value


def Writedata(File, Sheet, col, row, data):
    worksheet = openpyxl.load_workbook(File)
    Sheeet = worksheet[Sheet]
    Sheeet.cell(row=row, column=col).value = data


def Count(File, Sheet):
    worksheet = openpyxl.load_workbook(File)
    Sheeet = worksheet[Sheet]
    return Sheeet.max_row
